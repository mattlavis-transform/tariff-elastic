from openpyxl import load_workbook
import os
from dotenv import load_dotenv
import classes.globals as g
import ndjson


class ManualCategorisation:
    def __init__(self):
        load_dotenv('.env')
        self.manual_categorisation_file = os.getenv('MANUAL_CATEGORISATION_FILE')
        self.book = load_workbook(self.manual_categorisation_file, data_only=True)
        self.sheet = self.book.worksheets[0]
        g.filters = {}
        self.assignments = []
        
        row_count = self.sheet.max_row
        column_count = self.sheet.max_column
        
        row_count = 0
        for row in self.sheet.rows:
            row_count += 1
            if row_count == 1:
                self.get_column_headings(row)
            else:
                self.get_assignments(row)

        with open('assignments.ndjson', 'w') as f:
            ndjson.dump(self.assignments, f)

    def get_column_headings(self, row):
        # goods_nomenclature_item_id	type	productline_suffix	description	number_indents	Other?	synonym_lists	exclusion_lists
        column_count = 0
        for cell in row:
            column_count += 1
            if column_count > 9:
                filter = Filter(cell.value)
                filter_column = "column_" + str(column_count)
                g.filters[filter_column] = filter

    def get_assignments(self, row):
        assignment_set = AssignmentSet(row).assignments
        self.assignments += assignment_set


class Filter:
    def __init__(self, value):
        self.filter_title = value
        self.get_filter_name()

    def get_filter_name(self):
        self.filter_name = self.filter_title.replace(" ", "_").lower()


class AssignmentSet:
    def __init__(self, row):
        self.row = row
        self.assignments = []
        self.parse()
        
    def parse(self):
        self.goods_nomenclature_sid = self.row[0].value
        self.goods_nomenclature_item_id = self.row[1].value
        # self.entity_type = self.row[2].value
        self.productline_suffix = self.row[3].value
        column_count = 0
        for cell in self.row:
            column_count += 1
            if column_count > 9:
                value = cell.value
                if value != "" and value is not None:
                    assignment = Assignment(
                        self.goods_nomenclature_sid,
                        self.goods_nomenclature_item_id,
                        self.productline_suffix,
                        cell.value,
                        column_count
                    )
                    self.assignments.append(assignment.json)
            a = 1


class Assignment:
    def __init__(self, goods_nomenclature_sid, goods_nomenclature_item_id, productline_suffix, value, column_count):
        self.value = value
        self.goods_nomenclature_sid = goods_nomenclature_sid
        self.goods_nomenclature_item_id = goods_nomenclature_item_id
        self.productline_suffix = productline_suffix
        # self.entity_type = entity_type
        self.filter_name = g.filters["column_" + str(column_count)].filter_name
        self.value = value
        self.get_json()
        
    def get_json(self):
        self.json = {}
        self.json["goods_nomenclature_sid"] = self.goods_nomenclature_sid
        self.json["goods_nomenclature_item_id"] = self.goods_nomenclature_item_id
        self.json["productline_suffix"] = self.productline_suffix
        self.json["filter_name"] = "filter_" + self.filter_name
        self.json["value"] = self.value
