import os
import sys
import json
import ndjson
import openpyxl
import csv
from dotenv import load_dotenv
from openpyxl.cell import cell

from classes.commodity import Commodity
from classes.heading import Heading
from classes.facet import Facet


class Taxonomiser:
    def __init__(self):
        load_dotenv('.env')
        self.commodity_file = os.getenv('COMMODITY_MASTER')
        self.database_url = os.getenv('DATABASE_UK')
        self.facet_source_file = os.getenv('FACETS_MASTER')
        self.facet_source_file = os.path.join(os.getcwd(), "resources", "facets_source", self.facet_source_file)

        self.facet_export_file = os.getenv('FACETS_EXPORT_FILE')
        self.facets_export_folder = os.path.join(os.getcwd(), "resources", "facets_export")
        self.ndjson_file = os.path.join(self.facets_export_folder, self.facet_export_file)

        self.es_settings_template_path = os.getenv('ES_SETTINGS_TEMPLATE_PATH')
        self.es_settings_path = os.getenv('ES_SETTINGS_PATH')
        self.js_filter_path = os.getenv('JS_FILTER_PATH')

        self.commodities = []
        self.facets = {}
        self.headings = []
        self.headings_dict = {}

        self.commodity_minimum = os.getenv('COMMODITY_MINIMUM')
        self.commodity_maximum = os.getenv('COMMODITY_MAXIMUM')

        print("\nCreating file {filename}\n".format(filename=self.ndjson_file))

    def get_facets_from_excel_source(self):
        print("Getting facets from the source Excel spreadsheet")
        # Get the facets from the source Excel spreadsheet
        wb_obj = openpyxl.load_workbook(self.facet_source_file)
        sheet_facets = wb_obj["facets"]
        max_row = sheet_facets.max_row
        max_col = sheet_facets.max_column

        self.facets = []
        self.facets_dict = {}
        for i in range(2, max_row + 1):
            field = sheet_facets.cell(row=i, column=1).value
            display_name = sheet_facets.cell(row=i, column=2).value
            question = "What is the " + display_name.lower() + "?" if sheet_facets.cell(row=i, column=3).value is None else sheet_facets.cell(row=i, column=3).value
            info = sheet_facets.cell(row=i, column=4).value

            item = {
                "field": field,
                "display_name": display_name,
                "question": question,
                "info": info
            }
            self.facets.append(item)
            self.facets_dict[field] = item

        self.write_elasticsearch_indexing_json()

        # Get the headings
        sheet_headings = wb_obj["headings"]
        max_row = sheet_headings.max_row
        max_col = sheet_headings.max_column

        for i in range(2, max_row):
            h = Heading()
            h.id = sheet_headings.cell(row=i, column=1).value
            a = 1
            for j in range(3, max_col):
                v = sheet_headings.cell(row=i, column=j).value
                if v != "" and v is not None:
                    h.add_facet(v)

            self.headings.append(h)
            self.headings_dict[h.id] = h.facets
        wb_obj = None
        print("- Complete")

    def write_elasticsearch_indexing_json(self):
        # Read in the template
        f = open(self.es_settings_template_path, "r")
        data = json.load(f)
        f.close()
        mappings = data["mappings"]
        properties = mappings["properties"]

        js_filter_data = {}
        for facet in self.facets_dict:
            new_key = "filter_" + facet
            data["mappings"]["properties"][new_key] = {
                "type": "text",
                "fields": {
                        "raw": {
                            "type": "keyword"
                        }
                }
            }
            js_filter_data[facet] = {
                "display_name": self.facets_dict[facet]["display_name"],
                "question": self.facets_dict[facet]["question"],
                "info": self.facets_dict[facet]["info"]
            }

        # Write out the master file with the facets included
        json_object = json.dumps(data, indent=4)
        with open(self.es_settings_path, "w") as outfile:
            outfile.write(json_object)
        outfile.close()

        # Write out the JS filter path
        json_object = json.dumps(js_filter_data, indent=4)
        with open(self.js_filter_path, "w") as outfile:
            outfile.write(json_object)
        outfile.close()
        a = 1

    def get_commodities(self):
        # Loops through the list of commodities extracted from the "download Taric files > codes.py"
        # service: this list is then used to act as the base for adding filters to the extracted
        # data JSON file

        print("Taxonomiser - Getting commodities from CSV")

        with open(self.commodity_file) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=',')
            line_count = 0
            for row in csv_reader:
                if line_count == 0:
                    line_count += 1
                else:
                    if row[1] >= self.commodity_minimum and row[1] <= self.commodity_maximum:
                        c = Commodity(row)
                        self.commodities.append(c)
                        line_count += 1

        print("- Complete")

    def get_key_terms(self):
        print("Getting key terms")
        for c in self.commodities:
            if c.number_indents > -1:
                if c.heading in self.headings_dict:
                    c.expand_facets(self.headings_dict[c.heading])

        print("- Complete")

    def apply_commodity_hierarchy(self):
        print("Applying commodity hierarchy")
        # In which the hierarchy of the classification in formed by
        # looping from the end to the start and finding each decrement
        # of the number of indents

        # This code also inherits
        for i in range(len(self.commodities) - 1, -1, -1):
            c = self.commodities[i]
            indent_rolling = c.number_indents
            for j in range(i - 1, -1, -1):
                c2 = self.commodities[j]
                if c2.number_indents < indent_rolling:
                    c.hierarchy.append(c2)
                    indent_rolling = c2.number_indents
                if c2.number_indents == 0:
                    break

        for c in self.commodities:
            if c.goods_nomenclature_item_id == "8528420000":
                a = 1
            c.inherit_facets()
        print("- Complete")

    def save(self):
        print("Saving facets file")
        # Write a test JSON to validate that the assignment of key terms is working
        self.json = {}
        self.filters = []
        for c in self.commodities:
            self.json[c.goods_nomenclature_item_id] = c.facets
            for facet in c.facets:
                if len(c.facets[facet]) > 0:
                    assignment = {}
                    assignment["goods_nomenclature_sid"] = c.sid
                    assignment["goods_nomenclature_item_id"] = c.goods_nomenclature_item_id
                    assignment["productline_suffix"] = c.productline_suffix
                    assignment["filter_name"] = "filter_" + facet
                    assignment["value"] = c.facets[facet]

                    self.filters.append(assignment)

        # Write NDJSON file
        # This is used in emulate
        ndjson_file = os.path.join(self.facets_export_folder, self.facet_export_file)
        with open(ndjson_file, 'w') as f:
            ndjson.dump(self.filters, f)
        print("- Complete")
